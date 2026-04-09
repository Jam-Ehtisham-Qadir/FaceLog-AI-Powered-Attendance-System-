import os
import csv
import tempfile
import shutil
import calendar as cal_module
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from django.core.files.base import ContentFile
from .models import Employee, EmployeePhoto, AttendanceRecord
from .serializers import EmployeeSerializer, EmployeePhotoSerializer, AttendanceRecordSerializer
from .services import check_liveness, match_face, verify_attendance

# Pre-warm DeepFace on startup
try:
    from deepface import DeepFace
    import numpy as np
    dummy = np.zeros((100, 100, 3), dtype=np.uint8)
    DeepFace.represent(dummy, model_name="Facenet", enforce_detection=False)
except Exception:
    pass

ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'Test1234'

class AdminLoginView(APIView):
    def post(self, request):
        username = request.data.get('username', '')
        password = request.data.get('password', '')
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            return Response({"success": True, "token": "attendai-admin-token"})
        return Response({"success": False, "message": "Invalid credentials"}, status=status.HTTP_401_UNAUTHORIZED)

class AdminLogoutView(APIView):
    def post(self, request):
        return Response({"success": True})

class RegisterEmployeeView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        # Step 0: Check duplicate employee_id first (fast, no DeepFace needed)
        employee_id = request.data.get('employee_id', '').strip()
        if employee_id and Employee.objects.filter(employee_id=employee_id).exists():
            return Response({
                "success": False,
                "error": "duplicate_id",
                "message": f"Employee ID '{employee_id}' is already taken. Please use a different ID.",
            }, status=status.HTTP_409_CONFLICT)

        # Step 1: Check if a photo was provided for duplicate face detection
        primary_photo = request.FILES.get('photo')

        if primary_photo and Employee.objects.exists():
            temp_dir = tempfile.mkdtemp()
            temp_path = os.path.join(temp_dir, 'new_employee.jpg')
            try:
                with open(temp_path, 'wb') as f:
                    for chunk in primary_photo.chunks():
                        f.write(chunk)

                existing_employees = Employee.objects.all()
                matched_employee, confidence = match_face(temp_path, existing_employees)

                if matched_employee:
                    return Response({
                        "success": False,
                        "error": "duplicate_face",
                        "message": f"This face is already registered as '{matched_employee.name}' (ID: {matched_employee.employee_id}). Please delete the existing record first if you want to re-register.",
                        "existing_employee": {
                            "name": matched_employee.name,
                            "employee_id": matched_employee.employee_id,
                            "department": matched_employee.department,
                        },
                        "confidence": round(confidence * 100, 1)
                    }, status=status.HTTP_409_CONFLICT)

            except Exception as e:
                print(f"Duplicate face check error: {e}")
            finally:
                primary_photo.seek(0)
                shutil.rmtree(temp_dir, ignore_errors=True)

        # Step 2: No duplicates found — proceed with registration
        serializer = EmployeeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Employee registered successfully",
                "employee": serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response({"success": False, "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        employees = Employee.objects.all().order_by('-created_at')
        serializer = EmployeeSerializer(employees, many=True, context={'request': request})
        return Response(serializer.data)

class EmployeeDetailView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def delete(self, request, pk):
        try:
            employee = Employee.objects.get(pk=pk)
            employee.delete()
            return Response({"success": True, "message": "Employee deleted successfully"})
        except Employee.DoesNotExist:
            return Response({"success": False, "error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)


class AddEmployeePhotoView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, pk):
        try:
            employee = Employee.objects.get(pk=pk)
        except Employee.DoesNotExist:
            return Response({"success": False, "error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)

        photos = request.FILES.getlist('photos')
        if not photos:
            return Response({"success": False, "error": "No photos provided"}, status=status.HTTP_400_BAD_REQUEST)

        for photo in photos:
            EmployeePhoto.objects.create(employee=employee, photo=photo)

        return Response({
            "success": True,
            "message": f"{len(photos)} photo(s) added successfully",
            "total_photos": employee.extra_photos.count() + 1
        })


class CheckInView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        captured_photo = request.FILES.get('photo')
        if not captured_photo:
            return Response({"success": False, "error": "No photo provided"}, status=status.HTTP_400_BAD_REQUEST)

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, 'captured.jpg')
        with open(temp_path, 'wb') as f:
            for chunk in captured_photo.chunks():
                f.write(chunk)

        try:
            # Run anti-spoofing + face matching in parallel
            employees = Employee.objects.all()
            if not employees.exists():
                return Response({"success": False, "error": "No employees registered yet"}, status=status.HTTP_400_BAD_REQUEST)

            (is_live, spoof_reason, liveness_confidence, spoof_confidence), (matched_employee, match_confidence) = verify_attendance(temp_path, employees)

            if not is_live:
                record = AttendanceRecord.objects.create(
                    status='spoof',
                    confidence=liveness_confidence,
                    spoof_confidence=spoof_confidence,
                    spoof_reason=spoof_reason,
                    check_in_time=timezone.now()
                )
                captured_photo.seek(0)
                record.captured_photo.save(f'spoof_{record.id}.jpg', ContentFile(captured_photo.read()))
                return Response({
                    "success": False,
                    "status": "spoof",
                    "message": "Spoof attempt detected! Please show your real face.",
                    "reason": spoof_reason,
                    "liveness_confidence": liveness_confidence,
                    "spoof_confidence": spoof_confidence,
                })

            if matched_employee:
                today = timezone.now().date()

                # Duplicate check-in check
                already_checked_in = AttendanceRecord.objects.filter(
                    employee=matched_employee,
                    date=today,
                    status='checked_in'
                ).exists()

                if already_checked_in:
                    return Response({
                        "success": False,
                        "status": "duplicate",
                        "message": f"{matched_employee.name} has already checked in today.",
                        "employee": {
                            "name": matched_employee.name,
                            "employee_id": matched_employee.employee_id,
                            "department": matched_employee.department,
                        },
                        "confidence": match_confidence,
                    })

                now = timezone.now()
                record = AttendanceRecord.objects.create(
                    employee=matched_employee,
                    status='checked_in',
                    check_in_time=now,
                    confidence=match_confidence,
                    spoof_confidence=spoof_confidence,
                )
                captured_photo.seek(0)
                record.captured_photo.save(f'checkin_{record.id}.jpg', ContentFile(captured_photo.read()))

                return Response({
                    "success": True,
                    "status": "checked_in",
                    "message": f"Welcome, {matched_employee.name}! Check-in successful.",
                    "employee": {
                        "name": matched_employee.name,
                        "employee_id": matched_employee.employee_id,
                        "department": matched_employee.department,
                    },
                    "check_in_time": now.astimezone(pytz.timezone("Asia/Karachi")).strftime('%I:%M %p PKT'),
                    "confidence": match_confidence,
                    "liveness_confidence": liveness_confidence,
                    "spoof_confidence": spoof_confidence,
                })
            else:
                record = AttendanceRecord.objects.create(
                    status='unknown',
                    confidence=match_confidence,
                    spoof_confidence=spoof_confidence,
                    spoof_reason="Face not recognized - employee not registered",
                    check_in_time=timezone.now()
                )
                captured_photo.seek(0)
                record.captured_photo.save(f'unknown_{record.id}.jpg', ContentFile(captured_photo.read()))
                return Response({
                    "success": False,
                    "status": "not_registered",
                    "message": "Face not recognized. You are not registered in the system. Please contact HR.",
                    "confidence": match_confidence,
                    "liveness_confidence": liveness_confidence,
                })

        except Exception as e:
            return Response({"success": False, "error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            shutil.rmtree(temp_dir)


class CheckOutView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        captured_photo = request.FILES.get('photo')
        if not captured_photo:
            return Response({"success": False, "error": "No photo provided"}, status=status.HTTP_400_BAD_REQUEST)

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, 'captured.jpg')
        with open(temp_path, 'wb') as f:
            for chunk in captured_photo.chunks():
                f.write(chunk)

        try:
            employees = Employee.objects.all()
            if not employees.exists():
                return Response({"success": False, "error": "No employees registered yet"}, status=status.HTTP_400_BAD_REQUEST)

            (is_live, spoof_reason, liveness_confidence, spoof_confidence), (matched_employee, match_confidence) = verify_attendance(temp_path, employees)

            if not is_live:
                record = AttendanceRecord.objects.create(
                    status='spoof',
                    confidence=liveness_confidence,
                    spoof_confidence=spoof_confidence,
                    spoof_reason=spoof_reason,
                    check_out_time=timezone.now()
                )
                with open(temp_path, 'rb') as f:
                    record.captured_photo.save(f'spoof_{record.id}.jpg', ContentFile(f.read()))
                return Response({
                    "success": False,
                    "status": "spoof",
                    "message": "Spoof attempt detected! Please show your real face.",
                    "reason": spoof_reason,
                    "liveness_confidence": liveness_confidence,
                    "spoof_confidence": spoof_confidence,
                })

            if matched_employee:
                today = timezone.now().date()

                checkin_record = AttendanceRecord.objects.filter(
                    employee=matched_employee,
                    date=today,
                    status='checked_in'
                ).order_by('-check_in_time').first()

                if not checkin_record:
                    return Response({
                        "success": False,
                        "status": "not_checked_in",
                        "message": f"{matched_employee.name} has not checked in today. Please check in first.",
                        "employee": {
                            "name": matched_employee.name,
                            "employee_id": matched_employee.employee_id,
                        },
                    })

                now = timezone.now()
                hours_worked = round((now - checkin_record.check_in_time).total_seconds() / 3600, 2)

                checkin_record.status = 'checked_out'
                checkin_record.check_out_time = now
                checkin_record.hours_worked = hours_worked
                checkin_record.save()

                with open(temp_path, 'rb') as f:
                    checkin_record.captured_photo.save(f'checkout_{checkin_record.id}.jpg', ContentFile(f.read()))

                import pytz
                karachi = pytz.timezone('Asia/Karachi')
                checkin_local = checkin_record.check_in_time.astimezone(karachi)
                checkout_local = now.astimezone(karachi)

                return Response({
                    "success": True,
                    "status": "checked_out",
                    "message": f"Goodbye, {matched_employee.name}! See you tomorrow.",
                    "employee": {
                        "name": matched_employee.name,
                        "employee_id": matched_employee.employee_id,
                        "department": matched_employee.department,
                    },
                    "check_in_time": checkin_local.strftime('%I:%M %p PKT'),
                    "check_out_time": checkout_local.strftime('%I:%M %p PKT'),
                    "hours_worked": hours_worked,
                    "confidence": match_confidence,
                })
            else:
                record = AttendanceRecord.objects.create(
                    status='unknown',
                    confidence=match_confidence,
                    spoof_confidence=spoof_confidence,
                    spoof_reason="Face not recognized - employee not registered",
                    check_out_time=timezone.now()
                )
                with open(temp_path, 'rb') as f:
                    record.captured_photo.save(f'unknown_{record.id}.jpg', ContentFile(f.read()))
                return Response({
                    "success": False,
                    "status": "not_registered",
                    "message": "Face not recognized. You are not registered in the system. Please contact HR.",
                    "confidence": match_confidence,
                })

        except Exception as e:
            return Response({"success": False, "error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            shutil.rmtree(temp_dir)


class AttendanceRecordsView(APIView):
    def get(self, request):
        month = request.query_params.get('month')  # format: 2026-04
        page = int(request.query_params.get('page', 1))
        page_size = 20

        records = AttendanceRecord.objects.all().order_by('-date', '-check_in_time')

        if month:
            try:
                year, mon = month.split('-')
                records = records.filter(date__year=int(year), date__month=int(mon))
            except Exception:
                pass

        total = records.count()
        start = (page - 1) * page_size
        end = start + page_size
        records = records[start:end]

        serializer = AttendanceRecordSerializer(records, many=True)
        return Response({
            'records': serializer.data,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': max(1, (total + page_size - 1) // page_size),
        })


class ExportAttendanceView(APIView):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import pytz
        import calendar
        from datetime import date, timedelta

        month = request.query_params.get('month')
        karachi = pytz.timezone('Asia/Karachi')

        # Office rules
        LATE_CUTOFF_H, LATE_CUTOFF_M = 10, 15   # 10:15 AM
        HALF_DAY_H, HALF_DAY_M = 18, 0           # 6:00 PM

        records = AttendanceRecord.objects.filter(
            employee__isnull=False,
            status__in=['checked_in', 'checked_out']
        ).order_by('date', 'check_in_time')

        if month:
            try:
                year, mon = month.split('-')
                year, mon = int(year), int(mon)
                records = records.filter(date__year=year, date__month=mon)
                month_label = f"{year}-{str(mon).padStart(2,'0')}" if False else f"{year}-{str(mon).zfill(2)}"
                month_display = f"{calendar.month_name[mon]}, {year}"
                # All working days in this month (Mon-Sat)
                num_days = calendar.monthrange(year, mon)[1]
                all_working_days = [
                    date(year, mon, d)
                    for d in range(1, num_days + 1)
                    if date(year, mon, d).weekday() != 6  # 6 = Sunday
                ]
            except Exception:
                month_label = "All"
                all_working_days = []
        else:
            month_label = "All"
            all_working_days = []

        # Group by employee
        from collections import defaultdict
        employee_records = defaultdict(list)
        for record in records:
            employee_records[record.employee].append(record)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── Style helpers ──
        def hfont(bold=True, color="FFFFFF", size=11): return Font(bold=bold, color=color, size=size)
        def fill(hex): return PatternFill("solid", fgColor=hex)
        def bdr():
            t = Side(style='thin', color="CCCCCC")
            return Border(left=t, right=t, top=t, bottom=t)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_al = Alignment(horizontal='left', vertical='center')

        HDR_FILL   = fill("1e3a5f")
        SUB_FILL   = fill("2d4a7a")
        GREEN_FILL = fill("d1fae5")   # present on time
        RED_FILL   = fill("fee2e2")   # late
        ORANGE_FILL= fill("ffedd5")   # half day
        ABSENT_FILL= fill("f1f5f9")   # absent
        SUN_FILL   = fill("fef9c3")   # sunday
        YELLOW_FILL= fill("fefce8")   # checked in only

        for employee, emp_records in employee_records.items():

            sheet_name = employee.name[:31].replace('/', '-').replace('\\', '-').replace('*','').replace('?','').replace('[','').replace(']','')
            ws = wb.create_sheet(title=sheet_name)

            # ── Header info ──
            ws.merge_cells('A1:G1')
            month_display = f"{cal_module.month_name[mon]}, {year}" if all_working_days else month_label
            ws['A1'] = f"Attendance Report — {month_display}"
            ws['A1'].font = Font(bold=True, size=14, color="1e3a5f")
            ws['A1'].alignment = center

            info = [
                ("EMPLOYEE NAME:", employee.name),
                ("EMPLOYEE ID:",   employee.employee_id),
                ("DEPARTMENT:",    employee.department or '—'),
                ("PERIOD:",        month_label),
                ("PREPARED BY:",   "HR DEPARTMENT / AttendAI System"),
            ]
            for i, (k, v) in enumerate(info, start=2):
                ws.cell(row=i, column=1, value=k).font = Font(bold=True)
                ws.cell(row=i, column=2, value=v)

            # ── Build day-wise attendance map ──
            # Group records by date
            day_map = defaultdict(list)
            for r in emp_records:
                day_map[r.date].append(r)

            # Compute stats
            late_days = 0
            half_days = 0
            absent_days = 0
            present_days = 0
            total_hours = 0.0

            day_stats = {}  # date -> dict
            for wd in all_working_days:
                recs = day_map.get(wd, [])
                co_recs = [r for r in recs if r.status == 'checked_out']
                ci_recs = [r for r in recs if r.status == 'checked_in']

                if not recs:
                    absent_days += 1
                    day_stats[wd] = {'type': 'absent'}
                    continue

                present_days += 1

                # Use earliest check-in of the day
                all_recs = co_recs + ci_recs
                first_rec = min(all_recs, key=lambda r: r.check_in_time or r.check_out_time)
                # Use latest check-out of the day
                last_co = max(co_recs, key=lambda r: r.check_out_time) if co_recs else None

                ci_local = first_rec.check_in_time.astimezone(karachi) if first_rec.check_in_time else None
                co_local = last_co.check_out_time.astimezone(karachi) if last_co else None

                # Sum ALL hours worked across all cycles
                day_hours = sum(r.hours_worked or 0 for r in co_recs)

                is_late = ci_local and (ci_local.hour > LATE_CUTOFF_H or (ci_local.hour == LATE_CUTOFF_H and ci_local.minute > LATE_CUTOFF_M))
                is_half = co_local and (co_local.hour < HALF_DAY_H)
                not_checked_out = not co_recs and bool(ci_recs)

                if is_late: late_days += 1
                if is_half: half_days += 1
                total_hours += day_hours

                day_stats[wd] = {
                    'type': 'present',
                    'record': last_co or first_rec,
                    'ci_local': ci_local,
                    'co_local': co_local,
                    'hours': day_hours,
                    'is_late': is_late,
                    'is_half': is_half,
                    'not_checked_out': not_checked_out,
                }

            sundays = [
                date(year, mon, d)
                for d in range(1, calendar.monthrange(year, mon)[1] + 1)
                if date(year, mon, d).weekday() == 6
            ] if all_working_days else []

            total_calendar = calendar.monthrange(year, mon)[1] if all_working_days else 0

            # ── Summary table ──
            ws.merge_cells('A8:G8')
            ws['A8'] = "SUMMARY"
            ws['A8'].font = hfont(size=12)
            ws['A8'].fill = fill("374151")
            ws['A8'].alignment = center

            summary = [
                ("Total Calendar Days", total_calendar,        f"{calendar.month_name[mon] if all_working_days else '—'}"),
                ("Sundays (Holidays)", len(sundays),           "Weekly holiday"),
                ("Working Days",       len(all_working_days),  "Mon – Sat"),
                ("Days Present",       present_days,           ""),
                ("Days Absent",        absent_days,            "No attendance marked"),
                ("Late Days",          late_days,              "Check-in after 10:15 AM"),
                ("Half Days",          half_days,              "Check-out before 6:00 PM"),
                ("Total Hours Worked", f"{total_hours:.2f}h",  ""),
            ]

            sum_headers = ["Metric", "Count", "Notes"]
            for col, h in enumerate(sum_headers, start=1):
                c = ws.cell(row=9, column=col, value=h)
                c.font = hfont(); c.fill = SUB_FILL; c.alignment = center; c.border = bdr()

            for i, (metric, count, note) in enumerate(summary, start=10):
                row_fill = fill("fefce8") if i % 2 == 0 else fill("ffffff")
                for col, val in enumerate([metric, count, note], start=1):
                    c = ws.cell(row=i, column=col, value=val)
                    c.border = bdr()
                    c.fill = row_fill
                    c.alignment = center if col == 2 else left_al
                    if col == 1: c.font = Font(bold=True, size=10)

            ws.cell(row=18, column=1, value=f"Total working days present = {present_days}").font = Font(bold=True, size=11, color="1e3a5f")
            ws.merge_cells('A18:G18')

            # ── Daily records table ──
            TR = 20  # table row start
            day_headers = ["DATE", "DAY", "ATTENDANCE", "CHECK IN", "CHECK OUT", "HOURS WORKED", "NOTES"]
            for col, h in enumerate(day_headers, start=1):
                c = ws.cell(row=TR, column=col, value=h)
                c.font = hfont(); c.fill = HDR_FILL; c.alignment = center; c.border = bdr()

            # All calendar days (including Sundays)
            all_days = sorted(set(all_working_days) | set(sundays)) if all_working_days else sorted(day_map.keys())

            for row_i, d in enumerate(all_days, start=TR + 1):
                day_name = d.strftime('%A')
                is_sunday = d.weekday() == 6
                stat = day_stats.get(d)

                if is_sunday:
                    row_data = [str(d), day_name, "Sunday (Holiday)", "—", "—", "—", "Weekly Off"]
                    r_fill = SUN_FILL
                elif not stat:
                    row_data = [str(d), day_name, "ABSENT", "—", "—", "—", "No attendance"]
                    r_fill = ABSENT_FILL
                elif stat['type'] == 'absent':
                    row_data = [str(d), day_name, "ABSENT", "—", "—", "—", "No attendance"]
                    r_fill = ABSENT_FILL
                else:
                    rec = stat['record']
                    ci_str = stat['ci_local'].strftime('%I:%M %p') if stat['ci_local'] else '—'
                    co_str = stat['co_local'].strftime('%I:%M %p') if stat['co_local'] else '—'
                    hours = f"{stat['hours']:.2f}h" if stat.get('hours') else '—'
                    attendance = "PRESENT"
                    notes = []
                    if stat['is_late']: notes.append("Late arrival")
                    if stat['is_half']: notes.append("Left early")
                    if stat['not_checked_out']: notes.append("No check-out")

                    if stat['is_late'] and stat['is_half']:
                        r_fill = ORANGE_FILL
                        attendance = "HALF DAY / LATE"
                    elif stat['is_late']:
                        r_fill = RED_FILL
                        attendance = "LATE"
                    elif stat['is_half']:
                        r_fill = ORANGE_FILL
                        attendance = "HALF DAY"
                    elif stat['not_checked_out']:
                        r_fill = YELLOW_FILL
                        attendance = "CHECKED IN"
                    else:
                        r_fill = GREEN_FILL
                        attendance = "PRESENT"

                    row_data = [str(d), day_name, attendance, ci_str, co_str, hours, ", ".join(notes) or "On time"]

                for col, val in enumerate(row_data, start=1):
                    c = ws.cell(row=row_i, column=col, value=val)
                    c.fill = r_fill; c.alignment = center; c.border = bdr()
                    # Red font for late check-in time
                    if col == 4 and stat and stat.get('is_late'):
                        c.font = Font(color="DC2626", bold=True)

            # ── Column widths ──
            widths = [14, 12, 18, 12, 12, 14, 22]
            for col, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col)].width = w

        if not employee_records:
            ws = wb.create_sheet(title="No Data")
            ws['A1'] = f"No attendance records found for {month_label}"

        filename = f"attendance_{month_label}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class DashboardStatsView(APIView):
    def get(self, request):
        today = timezone.now().date()
        return Response({
            "total_employees": Employee.objects.count(),
            "checked_in_today": AttendanceRecord.objects.filter(date=today, status='checked_in').count(),
            "checked_out_today": AttendanceRecord.objects.filter(date=today, status='checked_out').count(),
            "spoof_attempts_today": AttendanceRecord.objects.filter(date=today, status='spoof').count(),
            "total_records": AttendanceRecord.objects.count(),
        })

class AttendanceRecordDeleteView(APIView):
    def delete(self, request, pk):
        try:
            record = AttendanceRecord.objects.get(pk=pk)
            record.delete()
            return Response({"success": True})
        except AttendanceRecord.DoesNotExist:
            return Response({"success": False, "error": "Record not found"}, status=status.HTTP_404_NOT_FOUND)

class AttendanceBulkDeleteView(APIView):
    def post(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"success": False, "error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count, _ = AttendanceRecord.objects.filter(id__in=ids).delete()
        return Response({"success": True, "deleted": deleted_count})
    
class ExportPayrollView(APIView):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import pytz
        import calendar
        from datetime import date
        from collections import defaultdict

        month = request.query_params.get('month')
        karachi = pytz.timezone('Asia/Karachi')

        LATE_CUTOFF_H, LATE_CUTOFF_M = 10, 15
        HALF_DAY_H = 18
        LATE_PENALTY = 500  # PKR per late day

        if not month:
            return Response({"error": "month parameter required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            year, mon = month.split('-')
            year, mon = int(year), int(mon)
            month_display = f"{calendar.month_name[mon]}, {year}"
            num_days = calendar.monthrange(year, mon)[1]
            all_working_days = [
                date(year, mon, d)
                for d in range(1, num_days + 1)
                if date(year, mon, d).weekday() != 6
            ]
            total_working_days = len(all_working_days)
        except Exception:
            return Response({"error": "Invalid month format. Use YYYY-MM"}, status=status.HTTP_400_BAD_REQUEST)

        records = AttendanceRecord.objects.filter(
            employee__isnull=False,
            status__in=['checked_in', 'checked_out'],
            date__year=year,
            date__month=mon
        ).order_by('date', 'check_in_time')

        # Group by employee
        employee_records = defaultdict(list)
        for r in records:
            employee_records[r.employee].append(r)

        # Include employees with no records (fully absent)
        all_employees = Employee.objects.all()
        for emp in all_employees:
            if emp not in employee_records:
                employee_records[emp] = []

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Style helpers
        def hfont(bold=True, color="FFFFFF", size=11):
            return Font(bold=bold, color=color, size=size)
        def fill(hex):
            return PatternFill("solid", fgColor=hex)
        def bdr():
            t = Side(style='thin', color="CCCCCC")
            return Border(left=t, right=t, top=t, bottom=t)
        center = Alignment(horizontal='center', vertical='center')
        left_al = Alignment(horizontal='left', vertical='center')
        right_al = Alignment(horizontal='right', vertical='center')

        HDR_FILL    = fill("1e3a5f")
        SUB_FILL    = fill("2d4a7a")
        GREEN_FILL  = fill("d1fae5")
        RED_FILL    = fill("fee2e2")
        YELLOW_FILL = fill("fef9c3")
        TOTAL_FILL  = fill("1e3a5f")

        for employee, emp_records in employee_records.items():
            sheet_name = employee.name[:31].replace('/', '-').replace('\\', '-').replace('*','').replace('?','').replace('[','').replace(']','')
            ws = wb.create_sheet(title=sheet_name)

            # ── Title ──
            ws.merge_cells('A1:F1')
            ws['A1'] = f"Payroll Report — {month_display}"
            ws['A1'].font = Font(bold=True, size=14, color="1e3a5f")
            ws['A1'].alignment = center

            # ── Employee info ──
            info = [
                ("EMPLOYEE NAME:", employee.name),
                ("EMPLOYEE ID:",   employee.employee_id),
                ("DEPARTMENT:",    employee.department or '—'),
                ("PERIOD:",        month_display),
                ("BASIC SALARY:",  f"PKR {float(employee.salary):,.0f}"),
                ("PREPARED BY:",   "HR DEPARTMENT / AttendAI System"),
            ]
            for i, (k, v) in enumerate(info, start=2):
                c = ws.cell(row=i, column=1, value=k)
                c.font = Font(bold=True)
                ws.cell(row=i, column=2, value=v)

            # ── Attendance computation ──
            day_map = defaultdict(list)
            for r in emp_records:
                day_map[r.date].append(r)

            present_days = 0
            absent_days = 0
            late_days = 0
            half_days = 0

            for wd in all_working_days:
                recs = day_map.get(wd, [])
                co_recs = [r for r in recs if r.status == 'checked_out']
                ci_recs = [r for r in recs if r.status == 'checked_in']
                all_r = co_recs + ci_recs

                if not all_r:
                    absent_days += 1
                    continue

                present_days += 1
                first_rec = min(all_r, key=lambda r: r.check_in_time or r.check_out_time)
                last_co = max(co_recs, key=lambda r: r.check_out_time) if co_recs else None

                ci_local = first_rec.check_in_time.astimezone(karachi) if first_rec.check_in_time else None
                co_local = last_co.check_out_time.astimezone(karachi) if last_co else None

                is_late = ci_local and (ci_local.hour > LATE_CUTOFF_H or (ci_local.hour == LATE_CUTOFF_H and ci_local.minute > LATE_CUTOFF_M))
                is_half = co_local and co_local.hour < HALF_DAY_H

                if is_late: late_days += 1
                if is_half: half_days += 1

            # ── Payroll calculation ──
            basic_salary = float(employee.salary)
            per_day_rate = basic_salary / 30  # Fixed 30-day month calculation

            paid_absent = 1  # 1 paid absent per month
            billable_absents = max(0, absent_days - paid_absent)
            absent_deduction = billable_absents * per_day_rate
            late_deduction    = late_days * LATE_PENALTY
            half_day_deduction = half_days * (per_day_rate * 0.5)
            total_deductions  = absent_deduction + late_deduction + half_day_deduction
            net_payable       = basic_salary - total_deductions

            # ── Summary section ──
            ws.merge_cells('A9:F9')
            ws['A9'] = "ATTENDANCE SUMMARY"
            ws['A9'].font = hfont(size=12)
            ws['A9'].fill = fill("374151")
            ws['A9'].alignment = center

            att_headers = ["Total Working Days", "Days Present", "Days Absent", "Late Days", "Half Days"]
            att_values  = [total_working_days, present_days, absent_days, late_days, half_days]

            for col, h in enumerate(att_headers, start=1):
                c = ws.cell(row=10, column=col, value=h)
                c.font = hfont(); c.fill = SUB_FILL; c.alignment = center; c.border = bdr()
            for col, v in enumerate(att_values, start=1):
                c = ws.cell(row=11, column=col, value=v)
                c.alignment = center; c.border = bdr(); c.font = Font(bold=True, size=11)
                if col == 3 and v > 0: c.fill = RED_FILL
                elif col == 4 and v > 0: c.fill = YELLOW_FILL
                elif col == 5 and v > 0: c.fill = YELLOW_FILL
                else: c.fill = GREEN_FILL

            # ── Payroll breakdown ──
            ws.merge_cells('A13:F13')
            ws['A13'] = "PAYROLL BREAKDOWN"
            ws['A13'].font = hfont(size=12)
            ws['A13'].fill = fill("374151")
            ws['A13'].alignment = center

            pay_headers = ["Component", "Details", "Amount (PKR)"]
            for col, h in enumerate(pay_headers, start=1):
                c = ws.cell(row=14, column=col, value=h)
                c.font = hfont(); c.fill = HDR_FILL; c.alignment = center; c.border = bdr()

            breakdown = [
                ("Basic Salary",        f"Fixed monthly salary",                                    basic_salary,       GREEN_FILL),
                ("Per Day Rate",        f"PKR {basic_salary:,.0f} ÷ 30 days",                      per_day_rate,       fill("f0f9ff")),
                ("Absent Deduction",    f"PKR {per_day_rate:,.0f} × {billable_absents} days ({absent_days} absent - 1 paid leave)",    -absent_deduction,  RED_FILL if billable_absents > 0 else fill("f8fafc")),
                ("Late Deduction",      f"PKR {LATE_PENALTY} × {late_days} late days",             -late_deduction,    YELLOW_FILL if late_days > 0 else fill("f8fafc")),
                ("Half Day Deduction",  f"PKR {per_day_rate/2:,.0f} × {half_days} half days",      -half_day_deduction,YELLOW_FILL if half_days > 0 else fill("f8fafc")),
                ("Total Deductions",    "",                                                          -total_deductions,  RED_FILL if total_deductions > 0 else fill("f8fafc")),
            ]

            for i, (component, details, amount, row_fill) in enumerate(breakdown, start=15):
                ws.cell(row=i, column=1, value=component).font = Font(bold=True)
                ws.cell(row=i, column=2, value=details).alignment = left_al
                amt_cell = ws.cell(row=i, column=3, value=f"PKR {amount:,.0f}")
                amt_cell.alignment = right_al
                amt_cell.font = Font(bold=True, color="DC2626" if amount < 0 and amount != 0 else "166534")
                for col in range(1, 4):
                    c = ws.cell(row=i, column=col)
                    c.fill = row_fill; c.border = bdr()

            # ── Net Payable ──
            net_row = 22
            ws.merge_cells(f'A{net_row}:B{net_row}')
            ws[f'A{net_row}'] = "NET PAYABLE SALARY"
            ws[f'A{net_row}'].font = Font(bold=True, size=13, color="FFFFFF")
            ws[f'A{net_row}'].fill = TOTAL_FILL
            ws[f'A{net_row}'].alignment = center

            ws[f'C{net_row}'] = f"PKR {net_payable:,.0f}"
            ws[f'C{net_row}'].font = Font(bold=True, size=13, color="FFFFFF")
            ws[f'C{net_row}'].fill = fill("10b981") if net_payable >= basic_salary * 0.5 else fill("ef4444")
            ws[f'C{net_row}'].alignment = center
            ws[f'C{net_row}'].border = bdr()

            # ── Column widths ──
            widths = [22, 36, 18, 14, 14, 14]
            for col, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col)].width = w

        filename = f"payroll_{year}-{str(mon).zfill(2)}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
class UpdateEmployeeSalaryView(APIView):
    def patch(self, request, pk):
        try:
            employee = Employee.objects.get(pk=pk)
            salary = request.data.get('salary')
            if salary is None:
                return Response({"success": False, "error": "salary required"}, status=status.HTTP_400_BAD_REQUEST)
            employee.salary = salary
            employee.save()
            return Response({"success": True, "salary": float(employee.salary)})
        except Employee.DoesNotExist:
            return Response({"success": False, "error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)